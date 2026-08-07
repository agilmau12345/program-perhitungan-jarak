import json
import os
import re
import time
import requests
from openpyxl import load_workbook
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError

FILE_EXCEL = "RFQ - FTL LH Tender 2026 - File 2 - CDD LONG (6WH).xlsx" #ubah nama file Excel sesuai kebutuhan
FILE_CHECKPOINT = "checkpoint_jarak.json"
USER_AGENT = "aplikasi_jarak_spx_v1"

KOLOM_ASAL_KOTA = "ORIGIN CITY"
KOLOM_ASAL_PROVINSI = "ORIGIN STATE"
KOLOM_TUJUAN_KOTA = "DESTINATION CITY"
KOLOM_TUJUAN_PROVINSI = "DESTINATION STATE"
KOLOM_JARAK = "Distance"

OSRM_SERVERS = [
    "https://routing.openstreetmap.de/routed-car/route/v1/driving",
    "https://router.project-osrm.org/route/v1/driving",
]

INTERVAL_CHECKPOINT = 100
INTERVAL_EXCEL = 100
INTERVAL_PROGRESS = 100

geolocator = Nominatim(user_agent=USER_AGENT)
cache_koordinat = {}
cache_jarak = {}


def bersihkan_nama(nama):
    nama = str(nama).strip().upper()

    for prefix in ["KAB. ", "KAB ", "KOTA ", "KOTA."]:
        if nama.startswith(prefix):
            nama = nama[len(prefix):]

    return nama.strip()


def bersihkan_provinsi(nama):
    nama = str(nama).strip()
    nama = re.sub(r"\(.*?\)", "", nama)
    return nama.strip()


def muat_checkpoint():
    global cache_koordinat, cache_jarak

    if not os.path.exists(FILE_CHECKPOINT):
        return

    try:
        with open(FILE_CHECKPOINT, "r", encoding="utf-8") as f:
            data = json.load(f)

        cache_koordinat = data.get("cache_koordinat", {})
        cache_jarak = data.get("cache_jarak", {})

        print(f"Checkpoint dimuat: {len(cache_koordinat)} koordinat dan {len(cache_jarak)} jarak sudah tersimpan.")
    except Exception:
        print("Checkpoint tidak bisa dibaca, mulai dari awal.")


def simpan_checkpoint():
    data = {
        "cache_koordinat": cache_koordinat,
        "cache_jarak": cache_jarak
    }

    with open(FILE_CHECKPOINT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def ambil_koordinat(kota, provinsi):
    kota = str(kota).strip()
    provinsi = str(provinsi).strip()
    kunci = f"{kota}|{provinsi}".lower()

    if not kota:
        return None, None

    if kunci in cache_koordinat:
        nilai = cache_koordinat[kunci]

        if nilai is None:
            return None, None

        return nilai[0], nilai[1]

    kota_bersih = bersihkan_nama(kota)
    provinsi_bersih = bersihkan_provinsi(provinsi)

    kandidat_kueri = []

    def tambah(kueri):
        if kueri and kueri not in kandidat_kueri:
            kandidat_kueri.append(kueri)

    if provinsi:
        tambah(f"{kota}, {provinsi}, Indonesia")
        tambah(f"{kota_bersih}, {provinsi}, Indonesia")

    if provinsi_bersih and provinsi_bersih != provinsi:
        tambah(f"{kota}, {provinsi_bersih}, Indonesia")
        tambah(f"{kota_bersih}, {provinsi_bersih}, Indonesia")

    tambah(f"{kota}, Indonesia")
    tambah(f"{kota_bersih}, Indonesia")
    tambah(kota_bersih)

    for kueri in kandidat_kueri:
        try:
            lokasi = geolocator.geocode(kueri, timeout=15)
            time.sleep(1.1)

            if lokasi:
                koordinat = [lokasi.latitude, lokasi.longitude]
                cache_koordinat[kunci] = koordinat
                return koordinat[0], koordinat[1]

        except (GeocoderTimedOut, GeocoderServiceError):
            time.sleep(1.1)
            continue
        except Exception:
            time.sleep(1.1)
            continue

    cache_koordinat[kunci] = None
    return None, None


def hitung_jarak_osrm(lat1, lon1, lat2, lon2):
    for server in OSRM_SERVERS:
        url = f"{server}/{lon1},{lat1};{lon2},{lat2}?overview=false"

        try:
            respons = requests.get(url, timeout=30)
            data = respons.json()

            if data.get("code") == "Ok":
                jarak_meter = data["routes"][0]["distance"]
                return jarak_meter / 1000

        except Exception:
            continue

    return None


def baca_header(ws):
    header = {}

    for nomor_kolom, sel in enumerate(ws[1], start=1):
        if sel.value is not None:
            header[str(sel.value).strip()] = nomor_kolom

    return header


def cari_kolom(header, nama):
    if nama in header:
        return header[nama]

    for kunci, nilai in header.items():
        if kunci.upper() == nama.upper():
            return nilai

    return None


def format_rentang(daftar_angka):
    if not daftar_angka:
        return ""

    angka = sorted(set(daftar_angka))
    hasil = []
    awal = angka[0]
    sebelumnya = angka[0]

    for n in angka[1:]:
        if n == sebelumnya + 1:
            sebelumnya = n
            continue

        if awal == sebelumnya:
            hasil.append(str(awal))
        else:
            hasil.append(f"{awal}-{sebelumnya}")

        awal = n
        sebelumnya = n

    if awal == sebelumnya:
        hasil.append(str(awal))
    else:
        hasil.append(f"{awal}-{sebelumnya}")

    return ", ".join(hasil)


def main():
    print("Memulai proses...")
    print("Jika program berhenti di tengah jalan, cukup jalankan ulang untuk melanjutkan.")

    muat_checkpoint()

    try:
        wb = load_workbook(FILE_EXCEL)
    except FileNotFoundError:
        print(f"File {FILE_EXCEL} tidak ditemukan.")
        return

    ws = wb.active
    header = baca_header(ws)

    kolom_asal_kota = cari_kolom(header, KOLOM_ASAL_KOTA)
    kolom_asal_provinsi = cari_kolom(header, KOLOM_ASAL_PROVINSI)
    kolom_tujuan_kota = cari_kolom(header, KOLOM_TUJUAN_KOTA)
    kolom_tujuan_provinsi = cari_kolom(header, KOLOM_TUJUAN_PROVINSI)
    kolom_jarak = cari_kolom(header, KOLOM_JARAK)

    kolom_wajib = {
        KOLOM_ASAL_KOTA: kolom_asal_kota,
        KOLOM_ASAL_PROVINSI: kolom_asal_provinsi,
        KOLOM_TUJUAN_KOTA: kolom_tujuan_kota,
        KOLOM_TUJUAN_PROVINSI: kolom_tujuan_provinsi,
        KOLOM_JARAK: kolom_jarak
    }

    for nama, posisi in kolom_wajib.items():
        if posisi is None:
            print(f"Kolom {nama} tidak ditemukan di Excel.")
            return

    total_baris = ws.max_row
    list_gagal = []
    list_terlewati = []
    jumlah_berhasil = 0

    sejak_checkpoint = 0
    sejak_excel = 0

    try:
        for row in range(2, total_baris + 1):
            if row % INTERVAL_PROGRESS == 0:
                print(f"Progres: baris {row} dari {total_baris}")

            try:
                asal_kota = ws.cell(row=row, column=kolom_asal_kota).value
                asal_provinsi = ws.cell(row=row, column=kolom_asal_provinsi).value
                tujuan_kota = ws.cell(row=row, column=kolom_tujuan_kota).value
                tujuan_provinsi = ws.cell(row=row, column=kolom_tujuan_provinsi).value
                jarak_terisi = ws.cell(row=row, column=kolom_jarak).value

                if jarak_terisi not in (None, ""):
                    list_terlewati.append(row)
                    continue

                if not asal_kota or not tujuan_kota:
                    list_terlewati.append(row)
                    continue

                lat_asal, lon_asal = ambil_koordinat(asal_kota, asal_provinsi or "")

                if lat_asal is None:
                    list_gagal.append(row)
                    continue

                lat_tujuan, lon_tujuan = ambil_koordinat(tujuan_kota, tujuan_provinsi or "")

                if lat_tujuan is None:
                    list_gagal.append(row)
                    continue

                kunci_rute = f"{asal_kota}|{asal_provinsi}>>{tujuan_kota}|{tujuan_provinsi}".lower()

                if kunci_rute in cache_jarak:
                    jarak_km = cache_jarak[kunci_rute]
                else:
                    jarak_km = hitung_jarak_osrm(lat_asal, lon_asal, lat_tujuan, lon_tujuan)

                    if jarak_km is not None:
                        cache_jarak[kunci_rute] = jarak_km
                        time.sleep(0.3)
                    else:
                        list_gagal.append(row)
                        continue

                ws.cell(row=row, column=kolom_jarak).value = round(jarak_km, 2)
                jumlah_berhasil += 1

                sejak_checkpoint += 1
                sejak_excel += 1

                if sejak_checkpoint >= INTERVAL_CHECKPOINT:
                    simpan_checkpoint()
                    sejak_checkpoint = 0

                if sejak_excel >= INTERVAL_EXCEL:
                    wb.save(FILE_EXCEL)
                    sejak_excel = 0

            except KeyboardInterrupt:
                raise
            except Exception:
                list_gagal.append(row)
                continue

    except KeyboardInterrupt:
        print("Program dihentikan manual. Progres disimpan.")
    finally:
        simpan_checkpoint()
        wb.save(FILE_EXCEL)

    print("")
    print("=" * 50)
    print("RINGKASAN PROSES")
    print("=" * 50)
    print(f"Berhasil diisi : {jumlah_berhasil} baris")
    print(f"Gagal          : {len(list_gagal)} baris")

    if list_gagal:
        print(f"  Baris ke     : {format_rentang(list_gagal)}")

    print(f"Terlewati      : {len(list_terlewati)} baris")

    if list_terlewati:
        print(f"  Baris ke     : {format_rentang(list_terlewati)}")

    print("=" * 50)
    print(f"File {FILE_EXCEL} telah diperbarui dan disimpan.")


if __name__ == "__main__":
    main()